# coding=utf-8
import os
import os.path
from openpyxl import load_workbook
from qgis.PyQt.QtCore import QFileInfo, \
    QSettings, Qt
from qgis.PyQt.QtSql import QSqlDatabase, QSqlQuery
from qgis.PyQt.QtWidgets import QApplication, QDialog, QDialogButtonBox, \
    QFileDialog
from qgis.core import QgsProject, \
    edit
from veriso.base.utils.exceptions import VerisoErrorWithBar
from veriso.base.utils.utils import (tr,
                                     get_ui_class)

FORM_CLASS = get_ui_class('importdefects.ui')


class ImportDefectsDialog(QDialog, FORM_CLASS):
    def __init__(self, application_module, iface, defects_list_dock, defects_type, defects_table_names):
        QDialog.__init__(self, None)
        self.setupUi(self)
        self.application_module = application_module
        self.iface = iface
        self.message_bar = self.iface.messageBar()
        self.defects_list_dock = defects_list_dock
        self.defects_type = defects_type
        self.defects_table_names = defects_table_names
        self.tr_tag = 'VeriBE (EE/EN)'  # TODO: translate in global context as well

        self.okButton = self.buttonBox.button(QDialogButtonBox.Ok)
        self.okButton.setText('Import')

        settings = QSettings("CatAIS", "VeriSO")
        self.input_xlsx_path = settings.value("project/projectdir")

        self.module_name = settings.value("project/appmodule")
        self.provider = settings.value("project/provider")
        self.db_host = settings.value("project/dbhost")
        self.db_port = settings.value("project/dbport")
        self.db_name = settings.value("project/dbname")
        self.db_schema = settings.value("project/dbschema")
        self.db_user = settings.value("project/dbuser")
        self.db_pwd = settings.value("project/dbpwd")

        self.db = None

        self.btnBrowseDefectsFile.clicked.connect(
            self.btnBrowseDefectsFile_clicked)

    def accept(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)

        file_to_import = self.lineEditDefectsFile.text().strip()

        if file_to_import == '':
            self.message_bar.pushWarning("VeriSO",
                                         tr("No Defects file set."))
            QApplication.restoreOverrideCursor()
            return

        extension = os.path.splitext(file_to_import)[1]
        if(extension == '.xlsx'):
            self.import_xlsx(file_to_import)
        elif(extension == '.shp'):
            self.import_shp(file_to_import)
        else:
            self.message_bar.pushWarning("VeriSO",
                                         tr("File must be .xlsx or .shp"))
            QApplication.restoreOverrideCursor()
            return

        QApplication.restoreOverrideCursor()
        self.close()

    def import_shp(self, shp):
        shp = self.lineEditDefectsFile.text().strip()
        lr = QgsProject.instance()
        error = False

        tmp_layer = self.iface.addVectorLayer(shp, 'tmp_imported_shp', 'ogr')
        self.application_module.do_load_defects_wrapper(self.defects_type)

        defect_layers = []

        if self.defects_table_names.get(self.defects_type, {}).get("point", None) is not None:
            defect_layers.append(lr.mapLayersByName(tr("Mängelliste (Punkte)", self.tr_tag))[0])
        if self.defects_table_names.get(self.defects_type, {}).get("line", None) is not None:
            defect_layers.append(lr.mapLayersByName(tr("Mängelliste (Linien)", self.tr_tag))[0])
        if self.defects_table_names.get(self.defects_type, {}).get("polygon", None) is not None:
            defect_layers.append(lr.mapLayersByName(tr("Mängelliste (Polygone)", self.tr_tag))[0])

        for feat in tmp_layer.getFeatures():
            feat.setAttribute('ogc_fid', None)

            # If the attribute does not exist, then we don't need to set anything
            try:
                if feat.attribute('erledigt') == 1 or feat.attribute(
                        'erledigt') is True:
                    feat.setAttribute('erledigt', True)
                else:
                    feat.setAttribute('erledigt', False)
            except KeyError:
                pass

            # If the attribute does not exist, then we don't need to set anything
            try:
                # Workaround for old exports with lowercase ja/nein
                if str(feat.attribute('verifikati')).lower() == 'ja':
                    feat.setAttribute('verifikati', 'Ja')
                if str(feat.attribute('verifikati')).lower() == 'nein':
                    feat.setAttribute('verifikati', 'Nein')
            except KeyError:
                pass

            try:
                with edit(defect_layers[tmp_layer.geometryType()]):
                    defect_layers[tmp_layer.geometryType()].addFeature(feat)
            except:
                error = True
                continue

        QgsProject.instance().removeMapLayers([tmp_layer.id()])

        if error:
            self.iface.messageBar().pushCritical(
                "VeriSo", "Not all features could be imported from Shapefile")
            return

        self.iface.messageBar().pushInfo(
            "VeriSo", "Defects imported from Shapefile")

    def import_xlsx(self, xlsx):  # NOTICE: does not work for translated files!

        self.wb = load_workbook(filename=xlsx, read_only=True)

        # Initialize rows lists
        rows_list_points = []
        rows_list_lines = []
        rows_list_polygons = []

        point_layer_name = self.defects_table_names.get(self.defects_type, None).get("point", None)
        line_layer_name = self.defects_table_names.get(self.defects_type, None).get("line", None)
        polygon_layer_name = self.defects_table_names.get(self.defects_type, None).get("polygon", None)

        if point_layer_name is not None:
            header_list_points, rows_list_points = self.read_sheet(
                tr('Mängelliste (Punkte)', self.tr_tag))
        if line_layer_name is not None:
            header_list_lines, rows_list_lines = self.read_sheet(
                tr('Mängelliste (Linien)', self.tr_tag))
        if polygon_layer_name is not None:
            header_list_polygons, rows_list_polygons = self.read_sheet(
                tr('Mängelliste (Polygone)', self.tr_tag))

        self.open_db()

        if len(rows_list_points) > 0:
            query_points = self.create_query(point_layer_name, header_list_points, rows_list_points)
            self.execute_query(query_points)
        if len(rows_list_lines) > 0:
            query_lines = self.create_query(line_layer_name, header_list_lines, rows_list_lines)
            self.execute_query(query_lines)
        if len(rows_list_polygons) > 0:
            query_polygons = self.create_query(polygon_layer_name, header_list_polygons, rows_list_polygons)
            self.execute_query(query_polygons)

        self.db.close()
        self.iface.messageBar().pushInfo("VeriSo",
                                         "Defects imported from Excel file")

        self.application_module.do_load_defects_wrapper(self.defects_type)

    def read_sheet(self, sheet_name):
        from builtins import str
        sheet = self.wb[sheet_name]

        # list of lists. A list per row
        rows_list = []
        header_list = []

        # check if the cells are in the original positions
        if sheet['A5'].value != 'ogc_fid':
            QApplication.restoreOverrideCursor()
            e = 'Wrong Excel file, cell A5 is not \'ogc_fid\''
            raise VerisoErrorWithBar(self.iface.messageBar(), e)
        else:
            # get the column names (row 5)

            for i in range(1, sheet.max_column + 1):
                header_list.append(sheet.cell(column=i, row=5).value)

            for row in range(6, sheet.max_row + 1):
                values_list = []

                for column in range(1, sheet.max_column + 1):
                    cell = sheet.cell(column=column, row=row)
                    if not cell.value:
                        values_list.append(None)
                    else:
                        values_list.append(str(cell.value))
                rows_list.append(values_list)

        return header_list, rows_list

    def create_query(self, table_name, header_list, rows_list):
        query = 'INSERT INTO ' + self.db_schema + '.' + table_name + '('
        # don't write ogc_fid and coordinates
        query += ', '.join(header_list[1:12])
        query += ', ' + 'the_geom)'
        query += ' VALUES '

        for i in rows_list:
            query += '('
            for j in range(1, 12):
                if not i[j] or i[j] == 'NULL':
                    query += 'NULL,'
                else:
                    query += '\''
                    query += i[j]
                    query += '\','
            query += 'ST_GeomFromText(\'' + i[-1] + '\', 2056)'
            query += '),'

        query = query[:-1]

        return query

    def open_db(self):

        try:
            self.db = QSqlDatabase.addDatabase("QPSQL")
            self.db.setHostName(self.db_host)
            self.db.setPort(int(self.db_port))
            self.db.setDatabaseName(self.db_name)
            self.db.setUserName(self.db_user)
            self.db.setPassword(self.db_pwd)
            self.db.open()

        except Exception as e:
            raise VerisoErrorWithBar(
                self.iface.messageBar(), "Error " + e.message)

    def execute_query(self, sql):
        query = QSqlQuery(self.db)

        res = query.exec_(sql)

        if res is False:
            QApplication.restoreOverrideCursor()
            raise VerisoErrorWithBar(self.iface.messageBar(), "Error " + (
                query.lastError().text()))

    # noinspection PyPep8Naming,PyPep8Naming
    def btnBrowseDefectsFile_clicked(self):
        file_path = QFileDialog.getOpenFileName(
            self,
            tr("Choose defects file"),
            self.input_xlsx_path,
            "Defects layer (*.shp)")[0]
        file_info = QFileInfo(file_path)
        self.lineEditDefectsFile.setText(file_info.absoluteFilePath())
